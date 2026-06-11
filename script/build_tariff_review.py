"""Tariff review: compare Ivory's per-code fees against the TRUE per-unit price in
Elixir (Amount / Quantity) over the last ~24 months, and flag where Ivory holds a
line-total (multi-unit) value instead of a unit fee. Outputs an Excel for Paul.

Usage: python script/build_tariff_review.py tmp/tariffs_raw.txt tmp/ivory_fees.json <out.xlsx>
"""
import json, re, sys
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MON = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
def dnum(s):
    m = re.match(r"(\d{1,2})-([A-Z]{3})-(\d{4})", s.strip())
    return int(m.group(3))*10000 + MON[m.group(2)]*100 + int(m.group(1)) if m else 0

raw, ivory_json, out = sys.argv[1], sys.argv[2], sys.argv[3]

# 1) Elixir billed lines per code
lines = defaultdict(list)  # code -> [(amount, qty, dnum, narr)]
for ln in open(raw, encoding="latin-1"):
    if "|@|" not in ln: continue
    p = [x.strip() for x in ln.rstrip("\n").split("|@|")]
    if len(p) < 5: continue
    code, narr, amt, dos, vat = p[:5]
    qty = p[5] if len(p) >= 6 else "1"
    if not re.match(r"^\d", code): continue            # skip P-CARD / payments
    u = narr.upper()
    if "PAYMENT" in u or "CREDIT CARD" in u or "RECEIVED" in u: continue
    try: amount = float(amt)
    except ValueError: continue
    if amount <= 0: continue
    try: q = int(float(qty))
    except ValueError: q = 1
    if q <= 0: q = 1
    lines[code].append((amount, q, dnum(dos), narr))

ivory = {x['code']: x for x in json.load(open(ivory_json, encoding="utf-8"))}
def mode(c): return c.most_common(1)[0][0] if c else None

rows = []
for code, recs in lines.items():
    n = len(recs)
    qtys = Counter(r[1] for r in recs)
    mode_qty = mode(qtys)
    pct_gt1 = round(100.0 * sum(1 for r in recs if r[1] > 1) / n)
    mode_unit = mode(Counter(round(r[0]/r[1], 2) for r in recs))
    mode_amount = mode(Counter(round(r[0], 2) for r in recs))
    latest = max(recs, key=lambda r: r[2])
    recent_unit = round(latest[0]/latest[1], 2)
    narr = Counter(r[3] for r in recs).most_common(1)[0][0]
    iv = ivory.get(code)
    iv_fee = iv['fee'] if iv else None
    multiples = (mode_qty and mode_qty > 1) or pct_gt1 >= 30
    if iv_fee is None:
        flag = "NOT IN IVORY — add it"
    elif multiples and abs(iv_fee - mode_amount) <= 0.5 and abs(mode_amount - recent_unit) > 0.5:
        flag = f"FIX — Ivory holds the x{mode_qty} line total, not the unit"
    elif abs(iv_fee - recent_unit) <= 0.5:
        flag = "OK — matches unit price"
    elif multiples:
        flag = "REVIEW — billed in multiples"
    else:
        flag = "REVIEW — fee differs"
    diff = round(iv_fee - recent_unit, 2) if iv_fee is not None else None
    rows.append(dict(code=code, desc=narr, n=n, mode_qty=mode_qty, pct_gt1=pct_gt1,
                     iv_fee=iv_fee, recent_unit=recent_unit, mode_unit=mode_unit,
                     line_total=mode_amount, diff=diff, proposed=recent_unit,
                     flag=flag, multiples=multiples))

billed = set(lines.keys())
for code, iv in ivory.items():
    if code not in billed and re.match(r"^\d", code):
        rows.append(dict(code=code, desc=iv['desc'], n=0, mode_qty=None, pct_gt1=None,
                         iv_fee=iv['fee'], recent_unit=None, mode_unit=None, line_total=None,
                         diff=None, proposed=None, flag="NO RECENT BILLING (24mo)", multiples=False))

def sortkey(r):
    pr = 0 if str(r['flag']).startswith("FIX") else (1 if r['multiples'] else (2 if str(r['flag']).startswith("REVIEW") else 3))
    return (pr, -(abs(r['diff']) if r['diff'] is not None else -1))
rows.sort(key=sortkey)

# ---- Excel ----
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tariff Review"
BLUE = PatternFill("solid", fgColor="4A73B5"); YEL = PatternFill("solid", fgColor="FFF3CD")
RED = PatternFill("solid", fgColor="F8D7DA"); GRN = PatternFill("solid", fgColor="D4EDDA")
GREY = PatternFill("solid", fgColor="EDEDED")
white_bold = Font(bold=True, color="FFFFFF"); bold = Font(bold=True)
thin = Side(style="thin", color="CCCCCC"); border = Border(left=thin, right=thin, top=thin, bottom=thin)

n_fix = sum(1 for r in rows if str(r['flag']).startswith("FIX"))
n_mult = sum(1 for r in rows if r['multiples'])
ws["A1"] = (f"Dr Chalita le Roux — Tariff fee review (Elixir last 24 months vs Ivory). "
            f"{n_mult} codes are billed in multiples; {n_fix} look like Ivory holds the multi-unit LINE TOTAL instead of the per-unit fee. "
            f"'PROPOSED unit fee' = most recent Elixir Amount÷Qty (VAT-incl). Review the highlighted rows and adjust Ivory.")
ws["A1"].font = Font(bold=True, size=11); ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
ws.merge_cells("A1:L1"); ws.row_dimensions[1].height = 54

headers = ["Code","Description","Times billed (24mo)","Usual qty","% lines qty>1",
           "Ivory current fee (R)","Elixir UNIT — most recent (R)","Elixir UNIT — most common (R)",
           "Elixir typical LINE TOTAL (R)","Ivory − unit (R)","PROPOSED Ivory unit fee (R)","Flag / Action"]
for j, h in enumerate(headers, 1):
    c = ws.cell(2, j, h); c.fill = BLUE; c.font = white_bold; c.border = border
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
ws.row_dimensions[2].height = 42

money = '#,##0.00'
for i, r in enumerate(rows, 3):
    vals = [r['code'], r['desc'], r['n'], r['mode_qty'], (f"{r['pct_gt1']}%" if r['pct_gt1'] is not None else ""),
            r['iv_fee'], r['recent_unit'], r['mode_unit'], r['line_total'], r['diff'], r['proposed'], r['flag']]
    for j, v in enumerate(vals, 1):
        c = ws.cell(i, j, v); c.border = border
        if j in (6,7,8,9,10,11) and isinstance(v, (int, float)): c.number_format = money
        if j == 12: c.alignment = Alignment(wrap_text=True, vertical="center")
    f = str(r['flag'])
    fill = RED if f.startswith("FIX") else YEL if (r['multiples'] or f.startswith("REVIEW")) else GRN if f.startswith("OK") else GREY
    for j in range(1, 13): ws.cell(i, j).fill = fill

widths = [9, 34, 13, 9, 10, 14, 15, 15, 16, 12, 16, 40]
for j, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:L{len(rows)+2}"
wb.save(out)
print(f"WROTE {out}: {len(rows)} codes ({n_mult} billed-in-multiples, {n_fix} likely line-total bugs)")
